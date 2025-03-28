import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def merge_tables_to_excel(
    head_table: pl.DataFrame,
    useful_table: pl.DataFrame,
    table_name: str,
    width: int = 8,
) -> Workbook:
    wb = Workbook()
    ws = wb.active

    # Получаем максимальную ширину второй таблицы
    max_columns = len(useful_table.columns)

    # Записываем первую таблицу
    # Объединяем ячейки для заголовка на всю ширину
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_columns)
    ws["A1"] = table_name
    ws["A1"].alignment = Alignment(horizontal="center")

    # Записываем данные первой таблицы, центрируя их
    start_row = 2
    for row in head_table.rows():
        # Объединяем ячейки для каждой строки первой таблицы
        print(row[0], row[1])
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=width // 2,
        )
        ws.merge_cells(
            start_row=start_row,
            start_column=width // 2 + 1,
            end_row=start_row,
            end_column=width,
        )

        ws.cell(row=start_row, column=1, value=str(row[0]))
        ws.cell(row=start_row, column=width // 2 + 1, value=str(row[1]))
        # ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center")
        # ws.cell(row=start_row, column=2).alignment = Alignment(horizontal="center")
        start_row += 1

    # Добавляем пустую строку между таблицами
    start_row += 1

    # Записываем вторую таблицу
    # Записываем заголовки
    for col_idx, col_name in enumerate(useful_table.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=col_name)

    # Записываем данные
    for row_idx, row in enumerate(useful_table.rows(), 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=start_row + row_idx, column=col_idx, value=value)

    # Автоподбор ширины столбцов
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                print(e)
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    return wb

