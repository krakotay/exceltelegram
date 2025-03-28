from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def format_excel(file_path: str, row_number: int):
    print("execute format_excel")
    # Загружаем существующий Excel файл
    wb = load_workbook(file_path)
    ws = wb.active

    # Автоподбор ширины столбцов для всего документа
    max_row = ws.max_row
    max_col = ws.max_column

    column_widths = {}

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)

            # Сохраняем ширину текста в ячейке
            if cell.value:
                column_widths[col] = max(
                    column_widths.get(col, 0), len(str(cell.value))
                )

            # Включаем перенос текста
            cell.alignment = Alignment(wrap_text=True)

    # Устанавливаем ширину столбцов
    for col, width in column_widths.items():
        if col == 1:
            ws.column_dimensions[
                get_column_letter(col)
            ].width = 5  # стандартная ширина * 3
        elif col == 2 or col == 3:
            ws.column_dimensions[
                get_column_letter(col)
            ].width = 15  # стандартная ширина * 3
        elif col == 4:
            ws.column_dimensions[
                get_column_letter(col)
            ].width = 30  # стандартная ширина * 3

        else:
            ws.column_dimensions[get_column_letter(col)].width = width / 2

    # Установка высоты строк начиная с row_number
    for row in range(row_number, max_row + 1):
        max_line_count = 0  # Отслеживаем максимальное количество строк в ячейке

        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                # Разбиваем текст на строки и определяем их количество
                line_count = len(str(cell.value).split("\n"))
                max_line_count = max(max_line_count, line_count)
        # Устанавливаем высоту строки, если текста достаточно
        if max_line_count > 1:
            ws.row_dimensions[row].height = (
                max_line_count * 15
            )  # кастомная высота на строку

    # Сохраняем изменения
    wb.save(file_path)
